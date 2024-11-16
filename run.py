from nis import cat
from openpyxl import load_workbook
import datetime
import pyexcel as p
import os

files = os.listdir('.')

for filename in files:
    # if file is not xlsx file, skip
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        continue

    # if file is xls open it and save it as xlsx
    if filename.endswith('.xls'):
        p.save_book_as(file_name=filename, dest_file_name=filename + 'x')

    print("Processing file: " + filename)

    w = load_workbook(filename)
    wb = w['Sheet']
    wb.insert_cols(1)
    wb.insert_cols(1)
    wb.delete_rows(1, 9)

    # search and delete last row
    temp_col = wb['C']
    wb.delete_rows(tuple(reversed(temp_col))[0].row, 1)

    colB = wb['C']

    company = None
    cui = None

    rows_to_be_deleted = []

    for cell in colB:
        if company == None and cell.value != None and not str(cell.value).startswith('Cont'):
            company = cell.value
            # print(wb[cell.row + 1][cell.column - 1].value)
            cod = wb[cell.row + 1][cell.column - 1].value
            if len(cod.split(' ')) > 5:
                cui = cod.split(' ')[4]
            else:
                cui = ''
        if str(cell.value).startswith('Total') or str(cell.value).startswith('Cont'):
            company = None
        if type(cell.value) == datetime.datetime:
            # append to the cell on the left of this cell, the company name
            row = cell.row
            col = cell.column - 1
            wb[row][col - 2].value = company
            wb[row][col - 1].value = cui
        else:
            rows_to_be_deleted.append(cell.row)

    # delete the rows
    for row in reversed(rows_to_be_deleted):
        # print(row)
        wb.delete_rows(row, 1)

    max_columns = wb.max_column

    # insert new column with formula
    for i, cell in enumerate(wb['K']):
        cell.value = '=TEXT(_xlfn.DAYS(D{0}, TODAY()), "0")'.format(i + 1)
        # cell_id = 'K{0}'.format(i + 1)
        # wb.formula_attributes[cell_id] = {'t': 'array', 'ref': '{0}:{0}'.format(cell_id)}

    # fix date format
    for i, cell in enumerate(wb['C']):
        cell.number_format = 'dd/mm/yyyy'
    for i, cell in enumerate(wb['D']):
        cell.number_format = 'dd/mm/yyyy'

    wb.column_dimensions['C'].width = 11
    wb.column_dimensions['D'].width = 11

    w.save('new' + filename)

    print("done")